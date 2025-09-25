#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
style_dashboard_plus.py — polished, client-ready Excel dashboard
- Creates/updates Dashboard sheet (KPIs + charts)
- Styles Opportunities + adds data bars
- Builds an Action Center from Action_Queue with:
    * Status dropdown (todo/in progress/blocked/done)
    * Owner dropdown (you + client placeholders; editable)
    * Due date formatting & overdue highlight
    * Severity color bands
- Adds a hidden "Lists" sheet for validation sources
- Optional logo + brand color

Usage:
  python .\\scripts\\style_dashboard_plus.py --in-xlsx .\\data\\outputs\\phase4\\phase4_dashboard.xlsx
  python .\\scripts\\style_dashboard_plus.py --in-xlsx .\\data\\outputs\\phase4\\phase4_dashboard.xlsx --out-xlsx .\\data\\outputs\\phase4\\phase4_dashboard.styled.xlsx --logo .\\data\\inputs\\logo.png --brand-color "#3B82F6"
"""
import argparse, re, datetime as dt
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import DataBarRule, FormulaRule

def hex_to_rgb(hex_color: str, default="3B82F6"):
    m = re.match(r"#?([0-9A-Fa-f]{6})", str(hex_color or default).strip())
    return m.group(1).upper() if m else default.upper()

def read_sheet(path: Path, name: str) -> pd.DataFrame:
    try:
        return pd.read_excel(path, sheet_name=name)
    except Exception:
        return pd.DataFrame()

def write_table(ws, df: pd.DataFrame, start_row=1, start_col=1, header=True):
    r = start_row
    c = start_col
    if header:
        for j, col in enumerate(df.columns, start=c):
            cell = ws.cell(row=r, column=j, value=str(col))
            cell.font = Font(bold=True)
        r += 1
    for _, row in df.iterrows():
        for j, col in enumerate(df.columns, start=c):
            ws.cell(row=r, column=j, value=row[col])
        r += 1
    return (start_row, start_col, r-1, start_col + len(df.columns) - 1)

def auto_width(ws, max_width=60):
    widths = {}
    for row in ws.iter_rows(values_only=True):
        for i, val in enumerate(row, start=1):
            if val is None: continue
            l = len(str(val))
            widths[i] = max(widths.get(i, 10), min(l + 2, max_width))
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i)].width = w

def build_lists_sheet(wb):
    name = "_Lists"
    if name in wb.sheetnames:
        ws = wb[name]
        for row in ws["A1:Z100"]:
            for cell in row:
                cell.value = None
    else:
        ws = wb.create_sheet(name)
    ws.sheet_state = "hidden"
    ws["A1"].value = "Status"
    for i, v in enumerate(["todo","in progress","blocked","done"], start=2):
        ws.cell(row=i, column=1, value=v)
    ws["C1"].value = "Owner"
    for i, v in enumerate(["Drewy","Client Team","SEO Vendor","Dev Team"], start=2):
        ws.cell(row=i, column=3, value=v)
    ws["E1"].value = "SeverityOrder"
    for i, v in enumerate(["critical","high","medium","low","unknown"], start=2):
        ws.cell(row=i, column=5, value=v)
    return ws

def style_dashboard(wb, in_path: Path, logo_path: str, brand_hex: str):
    brand = hex_to_rgb(brand_hex)
    bg = PatternFill("solid", fgColor=f"FF{brand}")
    h1 = Font(name="Calibri", size=20, bold=True, color="FFFFFF")
    h2 = Font(name="Calibri", size=14, bold=True)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    opp = read_sheet(in_path, "Opportunities")
    execsum = read_sheet(in_path, "Executive_Summary")
    action = read_sheet(in_path, "Action_Queue")

    # compute severity counts
    sev_counts = pd.DataFrame(columns=["Severity","Count"])
    if not action.empty and "Severity" in action.columns:
        sev_counts = action.groupby("Severity", dropna=False).size().reset_index(name="Count")
        sev_counts["Severity"] = sev_counts["Severity"].fillna("Unknown").astype(str)

    top10 = pd.DataFrame(columns=["keyword","score"])
    if not opp.empty:
        cols = {c.lower(): c for c in opp.columns}
        kw = cols.get("keyword") or list(opp.columns)[0]
        sc = cols.get("score")
        if sc:
            top10 = opp[[kw, sc]].dropna(subset=[sc]).sort_values(sc, ascending=False).head(10)
            top10.columns = ["keyword","score"]

    # Replace Dashboard
    if "Dashboard" in wb.sheetnames:
        wb.remove(wb["Dashboard"])
    ws = wb.create_sheet("Dashboard", 0)

    # Banner
    ws.merge_cells("A1:L3")
    ws["A1"].fill = bg
    ws["A1"].font = h1
    ws["A1"].alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
    ws["A1"].value = "SEO Performance Dashboard"
    if logo_path:
        lp = Path(logo_path)
        if lp.exists():
            try:
                img = XLImage(str(lp))
                img.height = 48
                img.width = 48
                ws.add_image(img, "K1")
            except Exception:
                pass

    # KPI cards — IMPORTANT: only set value on the TOP-LEFT cell of merged ranges
    kpis = [
        ("Opportunities (scored)", "B5", "E8"),
        ("Indexed URLs (P2 URLs rows)", "G5", "J8"),
        ("Raw Issues rows (Audit + P1)", "B10", "E13"),
        ("Last Generated (UTC)", "G10", "J13"),
    ]
    kv = {}
    if not execsum.empty and "KPI" in execsum.columns and "Value" in execsum.columns:
        for _, r in execsum.iterrows():
            kv[str(r["KPI"]).strip()] = r["Value"]

    for name, tl, br in kpis:
        ws.merge_cells(f"{tl}:{br}")
        cell = ws[tl]
        cell.fill = PatternFill("solid", fgColor="FFF3F4F6")
        cell.border = border
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        # Render title + value in ONE top-left cell to avoid merged-cell write error
        v = kv.get(name, "")
        try:
            if pd.notna(v) and str(v).isdigit():
                v = int(v)
        except Exception:
            pass
        cell.value = f"{name}\n{v}"  # one cell; newline; single font
        # optional row height tweak
        ws.row_dimensions[cell.row].height = 28

    # Headers
    ws["A15"].value = "Top Opportunities"
    ws["A15"].font = h2
    ws["G15"].value = "Issues by Severity"
    ws["G15"].font = h2

    # Top Opportunities chart
    if not top10.empty:
        # write table
        ws.cell(row=17, column=1, value="keyword").font = Font(bold=True)
        ws.cell(row=17, column=2, value="score").font = Font(bold=True)
        for i, (_, r) in enumerate(top10.iterrows(), start=18):
            ws.cell(row=i, column=1, value=r["keyword"])
            ws.cell(row=i, column=2, value=float(r["score"]))
        chart = BarChart()
        chart.height = 8; chart.width = 18
        values = Reference(ws, min_col=2, min_row=17, max_row=17+len(top10))
        cats   = Reference(ws, min_col=1, min_row=18, max_row=17+len(top10))
        chart.add_data(values, titles_from_data=True); chart.set_categories(cats)
        chart.title = "Top 10 by Score"; chart.y_axis.title = "Score"; chart.x_axis.title = "Keyword"
        ws.add_chart(chart, "C17")

    # Severity pie
    if not sev_counts.empty:
        ws.cell(row=17, column=7, value="Severity").font = Font(bold=True)
        ws.cell(row=17, column=8, value="Count").font = Font(bold=True)
        for i, (_, r) in enumerate(sev_counts.iterrows(), start=18):
            ws.cell(row=i, column=7, value=str(r["Severity"]))
            ws.cell(row=i, column=8, value=int(r["Count"]))
        pie = PieChart(); pie.height = 8; pie.width = 12
        labels = Reference(ws, min_col=7, min_row=18, max_row=17+len(sev_counts))
        data   = Reference(ws, min_col=8, min_row=17, max_row=17+len(sev_counts))
        pie.add_data(data, titles_from_data=True); pie.set_categories(labels)
        pie.title = "Issues by Severity"; ws.add_chart(pie, "I17")

    # Quick links
    ws["A30"].value = "Go to Opportunities →"; ws["A30"].hyperlink = "#'Opportunities'!A1"; ws["A30"].style = "Hyperlink"
    ws["C30"].value = "Go to Executive Summary →"; ws["C30"].hyperlink = "#'Executive_Summary'!A1"; ws["C30"].style = "Hyperlink"
    ws["F30"].value = "Go to Action Center →"; ws["F30"].hyperlink = "#'Action_Center'!A1"; ws["F30"].style = "Hyperlink"

    # Set widths & freeze
    for col, width in zip("ABCDEFGHIJKL", [28,12,2,2,2,2,18,10,2,2,2,2]):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A4"
    return ws

def style_opportunities(wb):
    if "Opportunities" not in wb.sheetnames: return
    ws = wb["Opportunities"]
    # find score column in header row 1
    sc_col = None
    for c in range(1, ws.max_column+1):
        v = str(ws.cell(row=1, column=c).value or "").strip().lower()
        if v == "score": sc_col = c; break
    if sc_col:
        start = 2; end = ws.max_row
        col_letter = get_column_letter(sc_col)
        rng = f"{col_letter}{start}:{col_letter}{end}"
        from openpyxl.styles import Color
        rule = DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, showValue=True, color=Color("FF63C384"))
        ws.conditional_formatting.add(rng, rule)
        ws.column_dimensions[col_letter].width = 12
        ws.freeze_panes = "A2"

def build_action_center(wb):
    # Clone from Action_Queue and add validation, formatting
    if "Action_Queue" not in wb.sheetnames:
        return
    src = wb["Action_Queue"]
    rows = list(src.values)
    if not rows:
        return
    headers = [str(h) if h is not None else "" for h in rows[0]]
    data = rows[1:]
    import pandas as pd
    df = pd.DataFrame(data, columns=headers)

    for col in ["Owner","ETA","Status"]:
        if col not in df.columns:
            df[col] = ""

    if "Action_Center" in wb.sheetnames:
        wb.remove(wb["Action_Center"])
    ws = wb.create_sheet("Action_Center")

    # write table
    # header
    for j, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=j, value=str(col))
        cell.font = Font(bold=True)
    # data
    for r_i, (_, row) in enumerate(df.iterrows(), start=2):
        for j, col in enumerate(df.columns, start=1):
            ws.cell(row=r_i, column=j, value=row[col])

    # Find column indices
    header_map = {str(ws.cell(row=1, column=i).value).strip().lower(): i for i in range(1, ws.max_column+1)}
    col_issue = header_map.get("issue", 1)
    col_sev   = header_map.get("severity", None)
    col_owner = header_map.get("owner", None)
    col_eta   = header_map.get("eta", None)
    col_status= header_map.get("status", None)

    # Data validations from hidden _Lists
    lists_ws = build_lists_sheet(wb)

    # Status dropdown
    if col_status:
        dv = DataValidation(type="list", formula1=f"=_Lists!$A$2:$A$5", allow_blank=True, showErrorMessage=True)
        ws.add_data_validation(dv)
        dv.add(f"{get_column_letter(col_status)}2:{get_column_letter(col_status)}{ws.max_row}")

    # Owner dropdown
    if col_owner:
        dv2 = DataValidation(type="list", formula1=f"=_Lists!$C$2:$C$5", allow_blank=True, showErrorMessage=True)
        ws.add_data_validation(dv2)
        dv2.add(f"{get_column_letter(col_owner)}2:{get_column_letter(col_owner)}{ws.max_row}")

    # ETA: date validation (basic)
    if col_eta:
        dv3 = DataValidation(type="date", operator="greaterThan", formula1="DATE(2000,1,1)", allow_blank=True)
        ws.add_data_validation(dv3)
        dv3.add(f"{get_column_letter(col_eta)}2:{get_column_letter(col_eta)}{ws.max_row}")
        for r in range(2, ws.max_row+1):
            ws.cell(row=r, column=col_eta).number_format = "yyyy-mm-dd"

    # Conditional formatting: Severity bands
    if col_sev:
        sev_letter = get_column_letter(col_sev)
        ws.conditional_formatting.add(f"{sev_letter}2:{sev_letter}{ws.max_row}",
            FormulaRule(formula=[f'LOWER(${sev_letter}2)="critical"'], stopIfTrue=True,
                        fill=PatternFill("solid", fgColor="FFFDEAEA")))
        ws.conditional_formatting.add(f"{sev_letter}2:{sev_letter}{ws.max_row}",
            FormulaRule(formula=[f'LOWER(${sev_letter}2)="high"'], stopIfTrue=True,
                        fill=PatternFill("solid", fgColor="FFFFF3E0")))
        ws.conditional_formatting.add(f"{sev_letter}2:{sev_letter}{ws.max_row}",
            FormulaRule(formula=[f'LOWER(${sev_letter}2)="medium"'], stopIfTrue=True,
                        fill=PatternFill("solid", fgColor="FFFFFDE7")))
        ws.conditional_formatting.add(f"{sev_letter}2:{sev_letter}{ws.max_row}",
            FormulaRule(formula=[f'LOWER(${sev_letter}2)="low"'], stopIfTrue=True,
                        fill=PatternFill("solid", fgColor="FFE8F5E9")))

    # Conditional formatting: overdue (ETA < today) and not done
    if col_eta and col_status:
        eta_letter = get_column_letter(col_eta)
        status_letter = get_column_letter(col_status)
        formula = f'AND(ISNUMBER(${eta_letter}2), ${eta_letter}2<TODAY(), LOWER(${status_letter}2)<>"done")'
        ws.conditional_formatting.add(f"{eta_letter}2:{eta_letter}{ws.max_row}",
            FormulaRule(formula=[formula], stopIfTrue=False,
                        fill=PatternFill("solid", fgColor="FFFFEBEE")))

    ws.freeze_panes = "A2"
    auto_width(ws)

def main():
    ap = argparse.ArgumentParser(description="Polish Phase-4 workbook into a client-ready dashboard")
    ap.add_argument("--in-xlsx", required=True)
    ap.add_argument("--out-xlsx", default="")
    ap.add_argument("--logo", default="")
    ap.add_argument("--brand-color", default="#3B82F6")
    args = ap.parse_args()

    in_path = Path(args.in_xlsx)
    if not in_path.exists():
        raise SystemExit(f"[FATAL] Input workbook not found: {in_path}")

    out_path = Path(args.out_xlsx) if args.out_xlsx else in_path
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(in_path)
    style_dashboard(wb, in_path, args.logo, args.brand_color)
    style_opportunities(wb)
    build_action_center(wb)
    wb.save(out_path)
    print(f"[INFO] Styled dashboard written to: {out_path}")

if __name__ == "__main__":
    raise SystemExit(main())
